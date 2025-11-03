"""Sales reporting system generating CPI-focused Excel dashboards.

This module provides utilities for reading a source Excel workbook that
contains sales data, cleaning and transforming the information, and
exporting a multi-sheet Excel report with tables and charts that summarise
invoiced, not invoiced and won CPI metrics.

The implementation follows the specification in the associated prompt and
is intentionally self-contained so it can be executed as a script:

    python sales_reporting.py input.xlsx output.xlsx

"""

from __future__ import annotations

import argparse
import os
from datetime import datetime
from typing import Callable, Iterable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


REQUIRED_COLUMNS = [
    "Date of Request",
    "Date of Issue",
    "Date of Delivery",
    "Sales Man",
    "Customer Name",
    "Customer DO No",
    "Definition",
    "Sales Ticket Reference",
    "SO No",
    "PO No",
    "Amount",
    "Total Discount",
    "CPI",
    "CPS",
    "QI Forecast",
    "Delivery Note",
    "Invoiced",
    "Invoiced Amount",
]

DATE_COLUMNS = ["Date of Request", "Date of Issue", "Date of Delivery"]
CURRENCY_COLUMNS = ["Amount", "Total Discount", "CPI", "CPS", "Invoiced Amount"]

TURKISH_MONTHS = {
    1: "Ocak",
    2: "Şubat",
    3: "Mart",
    4: "Nisan",
    5: "Mayıs",
    6: "Haziran",
    7: "Temmuz",
    8: "Ağustos",
    9: "Eylül",
    10: "Ekim",
    11: "Kasım",
    12: "Aralık",
}


def read_and_clean_data(filepath: str) -> pd.DataFrame:
    """Read the Excel file and perform initial validation and cleaning."""

    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Girdi dosyası bulunamadı: {filepath}")

    try:
        df = pd.read_excel(filepath)
    except Exception as exc:  # pragma: no cover - defensive
        raise RuntimeError(f"Excel dosyası okunamadı: {exc}") from exc

    missing_columns = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing_columns:
        raise ValueError(
            "Eksik sütunlar tespit edildi: " + ", ".join(missing_columns)
        )

    for date_col in DATE_COLUMNS:
        df[date_col] = df[date_col].apply(parse_turkish_date)

    for currency_col in CURRENCY_COLUMNS:
        df[currency_col] = pd.to_numeric(
            df[currency_col].apply(clean_currency_value), errors="coerce"
        )

    df["QI Forecast"] = df["QI Forecast"].astype(str).str.strip().str.upper()
    df["Invoiced"] = df["Invoiced"].apply(normalise_boolean)

    df["Year"] = df["Date of Issue"].dt.year
    df["MonthNumber"] = df["Date of Issue"].dt.month
    df["MonthName"] = df["MonthNumber"].map(TURKISH_MONTHS)
    df["MonthYear"] = (
        df.apply(
            lambda row: f"{int(row['Year'])} {row['MonthName']}"
            if pd.notna(row["Year"]) and pd.notna(row["MonthName"])
            else None,
            axis=1,
        )
    )

    return df


def parse_turkish_date(value: object) -> Optional[pd.Timestamp]:
    """Parse Turkish-formatted dates like '31.12.2024'."""

    if pd.isna(value) or value == "":
        return pd.NaT

    if isinstance(value, datetime):
        return pd.Timestamp(value)

    str_value = str(value).strip()

    for fmt in ("%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return pd.to_datetime(datetime.strptime(str_value, fmt))
        except ValueError:
            continue

    try:
        return pd.to_datetime(str_value, dayfirst=True, errors="coerce")
    except Exception:  # pragma: no cover - defensive
        return pd.NaT


def clean_currency_value(value: object) -> Optional[float]:
    """Convert strings such as '€ 1.234,56' to floats."""

    if pd.isna(value) or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value)

    cleaned = str(value)
    cleaned = cleaned.replace("€", "").replace("TL", "")
    cleaned = cleaned.replace(" ", "")
    cleaned = cleaned.replace(".", "").replace(",", ".")

    try:
        return float(cleaned)
    except ValueError:
        return None


def clean_percentage_value(value: object) -> Optional[float]:
    """Convert percentage strings to floating ratios (e.g. '%50,0' -> 0.5)."""

    if pd.isna(value) or value == "":
        return None

    if isinstance(value, (int, float)):
        return float(value) / 100 if value > 1 else float(value)

    cleaned = str(value).strip().replace("%", "")
    cleaned = cleaned.replace(".", "").replace(",", ".")

    try:
        number = float(cleaned)
        return number / 100 if number > 1 else number
    except ValueError:
        return None


def normalise_boolean(value: object) -> bool:
    """Return True when value represents a positive boolean state."""

    if isinstance(value, bool):
        return value

    if pd.isna(value):
        return False

    str_value = str(value).strip().lower()
    return str_value in {"yes", "evet", "true", "1", "y", "invoiced"}


def filter_monthly_data(df: pd.DataFrame, column: str) -> pd.DataFrame:
    """Aggregate CPI values by month keeping chronological order."""

    grouped = (
        df.dropna(subset=["MonthYear", "MonthNumber"])
        .groupby(["Year", "MonthNumber", "MonthName", "MonthYear"], as_index=False)[
            column
        ]
        .sum()
    )

    grouped = grouped.sort_values(["Year", "MonthNumber"])
    grouped.rename(columns={column: "Toplam"}, inplace=True)
    return grouped


def pivot_salesman_monthly(df: pd.DataFrame, value_column: str) -> pd.DataFrame:
    """Return a pivot table of MonthYear vs Sales Man for the provided value."""

    pivot = (
        df.dropna(subset=["MonthYear"])
        .pivot_table(
            index="MonthYear",
            columns="Sales Man",
            values=value_column,
            aggfunc="sum",
            fill_value=0.0,
        )
        .sort_index(key=lambda idx: idx.map(_month_year_sort_key))
    )

    pivot.reset_index(inplace=True)
    return pivot


def _month_year_sort_key(label: str) -> tuple[int, int]:
    year_str, month_name = label.split(" ", 1)
    month_number = next((k for k, v in TURKISH_MONTHS.items() if v == month_name), 0)
    return int(year_str), month_number


def salesperson_totals(df: pd.DataFrame, value_column: str) -> pd.DataFrame:
    totals = (
        df.groupby("Sales Man", as_index=False)[value_column].sum().sort_values(
            value_column, ascending=False
        )
    )
    return totals


def add_chart(
    workbook_sheet,
    data_range,
    category_range,
    chart_type: str,
    title: str,
    style: int = 13,
):
    """Create a chart of type ``chart_type`` in the given sheet."""

    if chart_type == "line":
        chart = LineChart()
        chart.style = style
    else:
        chart = BarChart()
        chart.style = style

    chart.title = title
    chart.y_axis.title = "CPI"
    chart.x_axis.title = "Ay"
    chart.height = 12
    chart.width = 22

    chart.add_data(data_range, titles_from_data=True)
    chart.set_categories(category_range)
    chart.legend.position = "r"

    workbook_sheet.add_chart(chart, "H2")


def apply_table_formatting(sheet, start_row: int, start_col: int, end_row: int, end_col: int):
    """Apply header styling and borders to a rectangular table."""

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=start_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border
            if row > start_row and col == start_col:
                cell.alignment = Alignment(horizontal="left")


def write_dataframe(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    startrow: int = 0,
    startcol: int = 0,
    index: bool = False,
):
    df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, startcol=startcol, index=index)
    sheet = writer.sheets[sheet_name]
    end_row = startrow + len(df) + (0 if index else 1)
    end_col = startcol + len(df.columns)
    apply_table_formatting(sheet, startrow + 1, startcol + 1, end_row, end_col)


def generate_sales_report(
    input_file: str,
    output_file: str,
    progress_callback: Optional[Callable[[float, str], None]] = None,
) -> None:
    """Main orchestration entry-point for creating the Excel report."""

    def report_progress(step: int, total_steps: int, message: str) -> None:
        if progress_callback is not None:
            progress_callback(step / total_steps, message)
        print(message)

    total_steps = 6
    report_progress(1, total_steps, "Veri okunuyor...")
    df = read_and_clean_data(input_file)
    report_progress(2, total_steps, f"Toplam kayıt sayısı: {len(df)}")

    invoiced_df = df[df["Invoiced"] == True]
    not_invoiced_df = df[df["Invoiced"] == False]
    won_df = df[df["QI Forecast"].str.upper() == "YES"]

    invoiced_monthly = filter_monthly_data(invoiced_df, "CPI")
    not_invoiced_monthly = filter_monthly_data(not_invoiced_df, "CPI")
    won_monthly = filter_monthly_data(won_df, "CPI")

    invoiced_pivot = pivot_salesman_monthly(invoiced_df, "CPI")
    not_invoiced_pivot = pivot_salesman_monthly(not_invoiced_df, "CPI")
    won_pivot = pivot_salesman_monthly(won_df, "CPI")

    salesperson_summary = salesperson_totals(df, "CPI")
    invoiced_salesperson = salesperson_totals(invoiced_df, "CPI")
    not_invoiced_salesperson = salesperson_totals(not_invoiced_df, "CPI")
    won_salesperson = salesperson_totals(won_df, "CPI")

    summary_metrics = pd.DataFrame(
        [
            {"Metrik": "Toplam CPI", "Değer": df["CPI"].sum()},
            {"Metrik": "Faturalanan CPI", "Değer": invoiced_df["CPI"].sum()},
            {"Metrik": "Faturalanmayan CPI", "Değer": not_invoiced_df["CPI"].sum()},
            {"Metrik": "Kazanılan CPI", "Değer": won_df["CPI"].sum()},
            {"Metrik": "Satış Elemanı Sayısı", "Değer": df["Sales Man"].nunique()},
        ]
    )

    if df["MonthYear"].notna().any():
        month_counts = df.dropna(subset=["MonthYear"])["MonthYear"].nunique()
        avg_cpi = df["CPI"].sum() / month_counts if month_counts else 0
    else:
        avg_cpi = 0

    summary_metrics.loc[len(summary_metrics.index)] = {
        "Metrik": "Aylık Ortalama CPI",
        "Değer": avg_cpi,
    }

    report_progress(3, total_steps, "Excel sayfaları hazırlanıyor...")

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        write_dataframe(writer, summary_metrics, "Özet Dashboard")
        write_dataframe(
            writer,
            salesperson_summary,
            "Özet Dashboard",
            startrow=len(summary_metrics) + 3,
        )

        write_dataframe(writer, invoiced_pivot, "CPI Faturalanan Raporu")
        write_dataframe(
            writer,
            invoiced_salesperson,
            "CPI Faturalanan Raporu",
            startrow=len(invoiced_pivot) + 3,
        )

        write_dataframe(writer, not_invoiced_pivot, "CPI Faturalanmayan Raporu")
        write_dataframe(
            writer,
            not_invoiced_salesperson,
            "CPI Faturalanmayan Raporu",
            startrow=len(not_invoiced_pivot) + 3,
        )

        write_dataframe(writer, won_pivot, "CPI Kazanılan Raporu")
        write_dataframe(
            writer,
            won_salesperson,
            "CPI Kazanılan Raporu",
            startrow=len(won_pivot) + 3,
        )

        write_dataframe(writer, df, "Detay Veri", index=False)

    workbook = load_workbook(output_file)

    _add_category_chart(
        workbook, "CPI Faturalanan Raporu", len(invoiced_pivot.columns)
    )
    _add_category_chart(
        workbook, "CPI Faturalanmayan Raporu", len(not_invoiced_pivot.columns)
    )
    _add_category_chart(workbook, "CPI Kazanılan Raporu", len(won_pivot.columns))

    workbook.save(output_file)

    report_progress(4, total_steps, "Grafikler ekleniyor...")

    report_progress(5, total_steps, "Rapor oluşturuldu.")
    report_progress(6, total_steps, f"Rapor kaydedildi: {output_file}")
    print(f"Faturalanan kayıt sayısı: {len(invoiced_df)}")
    print(f"Faturalanmayan kayıt sayısı: {len(not_invoiced_df)}")
    print(f"QI Forecast = YES kayıt sayısı: {len(won_df)}")


def _add_category_chart(workbook, sheet_name: str, column_count: int) -> None:
    sheet = workbook[sheet_name]
    if column_count <= 1:
        return

    max_row = sheet.max_row
    max_col = column_count

    data = Reference(sheet, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
    categories = Reference(sheet, min_col=1, min_row=2, max_row=max_row)

    chart_type = "line" if column_count <= 3 else "bar"
    title = sheet_name.replace("Raporu", "Grafiği")
    add_chart(sheet, data, categories, chart_type=chart_type, title=title)


def parse_arguments(args: Optional[Iterable[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Satış raporlama aracı")
    parser.add_argument(
        "input_file",
        nargs="?",
        default="sales_data_master.xlsx",
        help="Kaynak Excel dosyası",
    )
    parser.add_argument(
        "output_file",
        nargs="?",
        default=f"sales_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
        help="Oluşturulacak rapor dosyası",
    )
    return parser.parse_args(args)


def main(cli_args: Optional[Iterable[str]] = None) -> None:
    args = parse_arguments(cli_args)
    generate_sales_report(args.input_file, args.output_file)


if __name__ == "__main__":
    main()

